# FILE: src/services/report_service.py (ĐÃ TỐI ƯU HIỆU NĂNG CACHE)

from datetime import datetime, timedelta
from collections import defaultdict
from src.utils.larkbase import larkbase_get_all, larkbase_search_records 
import logging
import json
import time

logger = logging.getLogger(__name__)

def _get_text_from_lark_field(field_value, default='Không rõ'):
    """
    Trích xuất giá trị text từ một trường của Larkbase có thể là list hoặc string.
    """
    if isinstance(field_value, list) and field_value:
        first_item = field_value[0]
        if isinstance(first_item, dict) and 'text' in first_item:
            return first_item.get('text', default).strip()
    elif isinstance(field_value, str):
        return field_value.strip()
    return str(field_value or default).strip()


class ReportService:
    def __init__(self, app_token, table_id):
        self.app_token = app_token
        self.table_id = table_id
        self._cache = {}
        self.CACHE_TTL = 600  # 10 phút

    # ✅ TỐI ƯU: Hàm helper để lấy và cache toàn bộ records một lần duy nhất
    def _get_all_records_with_cache(self):
        """
        Lấy toàn bộ bản ghi từ Larkbase và cache lại. 
        Đây là nguồn dữ liệu chung cho các hàm get_all_* khác.
        """
        cache_key = 'all_records_raw' # Đặt tên key khác để tránh nhầm lẫn
        cached_data = self._get_from_cache(cache_key)
        if cached_data is not None:
            return cached_data

        # Nếu không có trong cache, gọi API
        logger.info("PERFORMANCE: Calling larkbase_get_all() ONCE to fetch all records.")
        all_records = larkbase_get_all(self.app_token, self.table_id)
        self._set_to_cache(cache_key, all_records)
        return all_records

    def get_daily_report(self, user_id=None, start_date_str=None, end_date_str=None, employee_filter=None, from_depot_filter=None, to_depot_filter=None, transport_provider_filter=None):
        logger.info("🚀 Bắt đầu tạo báo cáo bằng Lark SDK...")
        logger.info(f"   - Filters: employee={employee_filter}, from_depot={from_depot_filter}, to_depot={to_depot_filter}, transport={transport_provider_filter}")
        try:
            conditions = []
            if employee_filter and employee_filter.strip(): conditions.append({"field_name": "ID người bàn giao", "operator": "is", "value": [employee_filter.strip()]})
            if from_depot_filter and from_depot_filter.strip(): conditions.append({"field_name": "Kho đi", "operator": "contains", "value": [from_depot_filter.strip()]})
            if to_depot_filter and to_depot_filter.strip(): conditions.append({"field_name": "Kho đến", "operator": "contains", "value": [to_depot_filter.strip()]})
            if transport_provider_filter and transport_provider_filter.strip(): conditions.append({"field_name": "Đơn vị vận chuyển", "operator": "contains", "value": [transport_provider_filter.strip()]})
            if start_date_str:
                clean_start_date = start_date_str.strip()
                start_date_obj = datetime.strptime(clean_start_date, '%Y-%m-%d')
                start_timestamp = int(start_date_obj.timestamp() * 1000)
                conditions.append({"field_name": "Ngày bàn giao", "operator": "isGreater", "value": ["ExactDate", str(start_timestamp)]})
            if end_date_str:
                clean_end_date = end_date_str.strip()
                end_date_obj = datetime.strptime(clean_end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
                end_timestamp = int(end_date_obj.timestamp() * 1000)
                conditions.append({"field_name": "Ngày bàn giao", "operator": "isLess", "value": ["ExactDate", str(end_timestamp)]})
            filtered_records_raw = larkbase_search_records(app_token=self.app_token, table_id=self.table_id, filter_conditions=conditions)
            if not filtered_records_raw: return self._empty_report_data()
            filtered_records = [record.get('fields', {}) for record in filtered_records_raw]
            return self._calculate_daily_statistics_grouped_by_date(filtered_records)
        except Exception as e:
            logger.error(f"❌ Lỗi nghiêm trọng trong quá trình tạo báo cáo: {e}", exc_info=True)
            return self._empty_report_data()

    def _get_from_cache(self, key):
        if key in self._cache:
            data, timestamp = self._cache[key]
            if time.time() - timestamp < self.CACHE_TTL:
                logger.info(f"CACHE HIT: Lấy dữ liệu '{key}' từ cache.")
                return data
        logger.info(f"CACHE MISS: Dữ liệu '{key}' không có trong cache hoặc đã hết hạn.")
        return None

    def _set_to_cache(self, key, data):
        logger.info(f"CACHE SET: Đang lưu dữ liệu '{key}' vào cache.")
        self._cache[key] = (data, time.time())

    # ✅ TỐI ƯU: Hàm này giờ chỉ xử lý dữ liệu, không gọi API trực tiếp
    def get_all_employees(self):
        cache_key = 'employees_list'
        cached_data = self._get_from_cache(cache_key)
        if cached_data is not None:
            return cached_data
        try:
            all_records = self._get_all_records_with_cache()
            employees = {}
            for record in all_records:
                fields = record.get('fields', {})
                emp_id = _get_text_from_lark_field(fields.get('ID người bàn giao'), '')
                emp_name = _get_text_from_lark_field(fields.get('Người bàn giao'), '')
                if emp_id and emp_name and emp_id != 'Không rõ' and emp_id not in employees:
                    employees[emp_id] = emp_name
            
            result = [{'id': emp_id, 'name': emp_name} for emp_id, emp_name in sorted(employees.items())]
            self._set_to_cache(cache_key, result)
            return result
        except Exception as e:
            logger.error(f"Error processing employees from cached records: {e}")
            return []

    # ✅ TỐI ƯU: Hàm này giờ chỉ xử lý dữ liệu, không gọi API trực tiếp
    def get_all_depots(self):
        cache_key = 'depots_list'
        cached_data = self._get_from_cache(cache_key)
        if cached_data is not None:
            return cached_data
        try:
            all_records = self._get_all_records_with_cache()
            depot_names = set()
            for record in all_records:
                fields = record.get('fields', {})
                from_name = _get_text_from_lark_field(fields.get('Kho đi'), '')
                if from_name and from_name != 'Không rõ':
                    depot_names.add(from_name)
                to_name = _get_text_from_lark_field(fields.get('Kho đến'), '')
                if to_name and to_name != 'Không rõ':
                    depot_names.add(to_name)

            result = [{'id': name, 'name': name} for name in sorted(list(depot_names))]
            self._set_to_cache(cache_key, result)
            return result
        except Exception as e:
            logger.error(f"Error processing depots from cached records: {e}")
            return []

    # ✅ TỐI ƯU: Hàm này giờ chỉ xử lý dữ liệu, không gọi API trực tiếp
    def get_all_transport_providers(self):
        cache_key = 'transport_providers_list'
        cached_data = self._get_from_cache(cache_key)
        if cached_data is not None:
            return cached_data
        try:
            all_records = self._get_all_records_with_cache()
            providers = set()
            for record in all_records:
                fields = record.get('fields', {})
                provider_name = _get_text_from_lark_field(fields.get('Đơn vị vận chuyển'), '')
                if provider_name and provider_name != 'Không rõ':
                    providers.add(provider_name)
            result = [{'id': provider, 'name': provider} for provider in sorted(list(providers))]
            self._set_to_cache(cache_key, result)
            return result
        except Exception as e:
            logger.error(f"Error processing transport providers from cached records: {e}")
            return []
    
    # ... Các hàm tính toán thống kê còn lại giữ nguyên ...
    def _calculate_daily_statistics_grouped_by_date(self, records):
        if not records: return self._empty_report_data()
        
        daily_groups = defaultdict(list)
        for fields in records:
            handover_date = fields.get('Ngày bàn giao')
            date_str = 'Unknown'
            if handover_date:
                try:
                    dt_obj = datetime.fromtimestamp(int(handover_date) / 1000)
                    date_str = dt_obj.strftime('%Y-%m-%d')
                except (ValueError, TypeError): date_str = 'Unknown'
            daily_groups[date_str].append(fields)
        
        daily_statistics, route_summary_by_date, transport_summary_by_date = {}, defaultdict(list), defaultdict(list)
        
        for date_str, date_records in daily_groups.items():
            grouped_records = self._group_records_by_group_id(date_records)
            route_transport_stats = defaultdict(lambda: {'count': 0, 'bags': 0, 'loads': 0})
            total_loads = 0
            
            for group_key, group_records in grouped_records.items():
                if group_key.startswith('group_'):
                    total_loads += self._process_grouped_records(group_records, route_transport_stats, total_loads)
                else:
                    for fields in group_records:
                        total_loads += self._process_single_record(fields, route_transport_stats)
            
            daily_route_summary = []
            for route_transport_key, stats in route_transport_stats.items():
                route_part, transport_part = route_transport_key.split('|', 1)
                daily_route_summary.append({
                    'date': date_str, 'route': route_part, 'transport_provider': transport_part,
                    'count': stats['count'], 'bags': stats['bags'], 'loads': stats['loads']
                })
            daily_route_summary.sort(key=lambda x: x['loads'], reverse=True)
            route_summary_by_date[date_str] = daily_route_summary
            
            transport_stats = defaultdict(lambda: {'count': 0, 'bags': 0, 'loads': 0, 'routes': set()})
            for item in daily_route_summary:
                provider, stats = item['transport_provider'], transport_stats[item['transport_provider']]
                stats['count'] += item['count']; stats['bags'] += item['bags']; stats['loads'] += item['loads']
                stats['routes'].add(item['route'])
            
            daily_transport_summary = []
            for provider, stats in transport_stats.items():
                daily_transport_summary.append({
                    'date': date_str, 'transport_provider': provider, 'count': stats['count'],
                    'bags': stats['bags'], 'loads': stats['loads'], 'route_count': len(stats['routes'])
                })
            daily_transport_summary.sort(key=lambda x: x['loads'], reverse=True)
            transport_summary_by_date[date_str] = daily_transport_summary
            
            daily_statistics[date_str] = {
                'total_records': len(date_records), 'total_quantity': total_loads,
                'route_summary': daily_route_summary, 'transport_summary': daily_transport_summary
            }
        
        all_route_summary, all_transport_summary = [], []
        for date_str in sorted(daily_statistics.keys()):
            all_route_summary.extend(route_summary_by_date[date_str])
            all_transport_summary.extend(transport_summary_by_date[date_str])
        
        return {
            'total_records': sum(s['total_records'] for s in daily_statistics.values()),
            'total_quantity': sum(s['total_quantity'] for s in daily_statistics.values()),
            'route_summary': all_route_summary, 'transport_summary': all_transport_summary,
            'daily_statistics': daily_statistics,
            'date_list': sorted([d for d in daily_statistics.keys() if d != 'Unknown'])
        }

    def _empty_report_data(self):
        return {'total_records': 0, 'total_quantity': 0, 'transport_providers': {}, 'routes': {}, 'transport_summary': [], 'route_summary': []}

    def _group_records_by_group_id(self, records):
        grouped, single_counter = {}, 0
        for fields in records:
            group_id = _get_text_from_lark_field(fields.get('Group ID'), None)
            if group_id:
                group_key = f"group_{group_id}"
                if group_key not in grouped: grouped[group_key] = []
                grouped[group_key].append(fields)
            else:
                single_counter += 1
                grouped[f"single_{single_counter}"] = [fields]
        return grouped

    def _process_grouped_records(self, group_records, route_transport_stats, total_loads):
        if not group_records: return 0
        first_record, total_bags = group_records[0], 0
        for fields in group_records:
            try: total_bags += int(fields.get('Số lượng túi', 0) or 0)
            except (ValueError, TypeError): continue
        try: loads = int(first_record.get('Số lượng bao', 0) or 0)
        except (ValueError, TypeError): loads = 0
        transport_provider = _get_text_from_lark_field(first_record.get('Đơn vị vận chuyển'))
        from_depot = _get_text_from_lark_field(first_record.get('Kho đi'))
        to_depot = _get_text_from_lark_field(first_record.get('Kho đến'))
        route_key = f"{from_depot} → {to_depot}"
        route_transport_key = f"{route_key}|{transport_provider}"
        stats = route_transport_stats[route_transport_key]
        stats['count'] += len(group_records); stats['bags'] += total_bags; stats['loads'] += loads
        return loads

    def _process_single_record(self, fields, route_transport_stats):
        try: bags = int(fields.get('Số lượng túi', 0) or 0)
        except (ValueError, TypeError): bags = 0
        try: loads = int(fields.get('Số lượng bao', 0) or 0)
        except (ValueError, TypeError): loads = 0
        transport_provider = _get_text_from_lark_field(fields.get('Đơn vị vận chuyển'))
        from_depot = _get_text_from_lark_field(fields.get('Kho đi'))
        to_depot = _get_text_from_lark_field(fields.get('Kho đến'))
        route_key = f"{from_depot} → {to_depot}"
        route_transport_key = f"{route_key}|{transport_provider}"
        stats = route_transport_stats[route_transport_key]
        stats['count'] += 1; stats['bags'] += bags; stats['loads'] += loads
        return loads

    def export_route_records_to_excel(self, from_depot, to_depot, start_date_str=None, end_date_str=None, employee_filter=None, transport_provider_filter=None):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            import io
            
            all_records = self._get_all_records_with_cache() # Sử dụng cache chung
            route_records = []
            
            for record in all_records:
                fields = record.get('fields', {})
                handover_date = fields.get('Ngày bàn giao')
                
                if start_date_str:
                    start_dt = datetime.strptime(start_date_str, '%Y-%m-%d')
                    if handover_date and datetime.fromtimestamp(handover_date/1000) < start_dt: continue
                if end_date_str:
                    end_dt = datetime.strptime(end_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
                    if handover_date and datetime.fromtimestamp(handover_date/1000) > end_dt: continue
                
                if employee_filter and _get_text_from_lark_field(fields.get('ID người bàn giao'), '') != employee_filter: continue
                record_transport = _get_text_from_lark_field(fields.get('Đơn vị vận chuyển'), '')
                if transport_provider_filter and transport_provider_filter.strip() and record_transport != transport_provider_filter: continue
                record_from_depot = _get_text_from_lark_field(fields.get('Kho đi'), '')
                record_to_depot = _get_text_from_lark_field(fields.get('Kho đến'), '')
                if record_from_depot == from_depot and record_to_depot == to_depot:
                    route_records.append(fields)
            
            if not route_records: return None, 0
            
            grouped_records = self._group_records_for_export(route_records)
            wb, ws = Workbook(), wb.active
            
            title_parts = [f"{from_depot} → {to_depot}"]
            if start_date_str and end_date_str and start_date_str != end_date_str: title_parts.append(f"({start_date_str} to {end_date_str})")
            elif start_date_str: title_parts.append(f"({start_date_str})")
            if transport_provider_filter: title_parts.append(f"[{transport_provider_filter}]")
            ws.title = " ".join(title_parts)[:31]
            
            header_font, header_fill = Font(bold=True, color="FFFFFF"), PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            headers = ["ID", "Số lượng túi", "Số lượng bao", "Số lượng sản phẩm yêu cầu", "ID người bàn giao", "Người bàn giao"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font, cell.fill = header_font, header_fill
            
            current_row = 2
            for group_key, group_records in grouped_records.items():
                is_grouped = len(group_records) > 1
                for i, record in enumerate(group_records):
                    ws.cell(row=current_row + i, column=1, value=_get_text_from_lark_field(record.get('ID'), ''))
                    ws.cell(row=current_row + i, column=2, value=record.get('Số lượng túi', 0))
                    ws.cell(row=current_row + i, column=4, value=record.get('Số lượng sản phẩm yêu cầu', 0))
                    ws.cell(row=current_row + i, column=5, value=_get_text_from_lark_field(record.get('ID người bàn giao'), ''))
                    ws.cell(row=current_row + i, column=6, value=_get_text_from_lark_field(record.get('Người bàn giao'), ''))
                
                if is_grouped:
                    first_record = group_records[0]
                    ws.cell(row=current_row, column=3, value=first_record.get('Số lượng bao', 0))
                    ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row + len(group_records) - 1, end_column=3)
                else:
                    ws.cell(row=current_row, column=3, value=group_records[0].get('Số lượng bao', 0))
                current_row += len(group_records)

            for column in ws.columns:
                max_length = 0
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
            
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            return excel_buffer, len(route_records)
            
        except Exception as e:
            logger.error(f"Error creating route Excel export: {e}", exc_info=True)
            return None, 0

    def _group_records_for_export(self, records):
        grouped, single_counter = {}, 0
        for record in records:
            group_id = _get_text_from_lark_field(record.get('Group ID'), None)
            if group_id:
                if group_id not in grouped: grouped[group_id] = []
                grouped[group_id].append(record)
            else:
                single_counter += 1
                grouped[f"single_{single_counter}"] = [record]
        return grouped